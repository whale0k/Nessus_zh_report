#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel generator module - Generate vulnerability remediation Excel reports
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
from datetime import datetime
from typing import List, Dict

from config.constants import AppConstants
from models import VulnerabilityInfo

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate Excel vulnerability remediation reports"""

    def __init__(self):
        self.header_font = Font(
            name='微软雅黑',
            size=12,
            bold=True
        )
        self.data_font = Font(
            name='微软雅黑',
            size=10
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @staticmethod
    def translate_risk_level(risk_level: str) -> str:
        """Translate English risk levels to Chinese"""
        return AppConstants.RISK_LEVEL_MAPPING.get(risk_level, risk_level)

    def create_remediation_report(
        self,
        vulnerabilities: List[VulnerabilityInfo],
        output_file: str = None,
        filter_level: str = "all"
    ) -> str:
        """Create vulnerability remediation Excel report"""
        if not output_file:
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"漏洞整改表({current_date}).xlsx"

        if filter_level == "high_above":
            high_risk_levels = ["Critical", "High"]
            filtered_vulnerabilities = [v for v in vulnerabilities if v.risk in high_risk_levels]
            logger.info(
                f"Filtered to {len(filtered_vulnerabilities)} high-risk and above "
                f"vulnerabilities (from {len(vulnerabilities)} total)"
            )
        else:
            filtered_vulnerabilities = vulnerabilities
            logger.info(f"Including all {len(vulnerabilities)} vulnerabilities")

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "漏洞整改表"

            headers = [
                "序号", "CVE编号", "漏洞等级", "主机", "端口", "协议",
                "漏洞名称", "漏洞详情", "修复建议", "下发时间", "整改情况"
            ]

            self._set_headers(ws, headers)

            current_date = datetime.now().strftime("%Y-%m-%d")
            self._add_vulnerability_data(ws, filtered_vulnerabilities, current_date)

            self._add_borders(ws, len(filtered_vulnerabilities), len(headers))

            self._set_row_heights(ws, len(filtered_vulnerabilities))

            wb.save(output_file)
            logger.info(f"Excel report saved: {output_file}")

            return output_file

        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            raise

    def _set_headers(self, ws, headers: List[str]) -> None:
        """Set headers and column widths"""
        width_map = {
            "漏洞详情": 50, "修复建议": 50,
            "漏洞名称": 30, "下发时间": 20, "整改情况": 20,
            "序号": 8
        }

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.fill = self.header_fill

            width = width_map.get(header, 15)
            ws.column_dimensions[cell.column_letter].width = width

    def _add_vulnerability_data(
        self,
        ws,
        vulnerabilities: List[VulnerabilityInfo],
        current_date: str
    ) -> None:
        """Add vulnerability data rows"""
        for row, vuln in enumerate(vulnerabilities, 2):
            sequence_number = row - 1
            data = [
                sequence_number,
                vuln.cve,
                self.translate_risk_level(vuln.risk),
                vuln.host,
                vuln.port,
                vuln.protocol,
                vuln.name,
                vuln.description,
                vuln.solution,
                current_date,
                ""
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.data_font
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                if col == 1 or col == 10:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    def _add_borders(self, ws, row_count: int, col_count: int) -> None:
        """Add borders to all cells"""
        for row in ws.iter_rows(
            min_row=1, max_row=row_count + 1,
            min_col=1, max_col=col_count
        ):
            for cell in row:
                cell.border = self.thin_border

    def _set_row_heights(self, ws, row_count: int) -> None:
        """Set row heights for data rows"""
        for row in range(2, row_count + 2):
            ws.row_dimensions[row].height = 60.0

    def create_summary_report(
        self,
        vulnerabilities: List[VulnerabilityInfo],
        host_risks: Dict,
        output_file: str = None
    ) -> str:
        """Create vulnerability summary Excel report"""
        if not output_file:
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"漏洞统计汇总({current_date}).xlsx"

        try:
            wb = openpyxl.Workbook()

            ws1 = wb.active
            ws1.title = "漏洞统计"
            self._create_vulnerability_stats_sheet(ws1, vulnerabilities)

            ws2 = wb.create_sheet("主机统计")
            self._create_host_stats_sheet(ws2, host_risks)

            for ws in [ws1, ws2]:
                self._auto_fit_columns(ws)

            wb.save(output_file)
            logger.info(f"Summary report saved: {output_file}")

            return output_file

        except Exception as e:
            logger.error(f"Error creating summary report: {e}")
            raise

    def _create_vulnerability_stats_sheet(
        self,
        ws,
        vulnerabilities: List[VulnerabilityInfo]
    ) -> None:
        """Create vulnerability statistics sheet"""
        risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
        for vuln in vulnerabilities:
            risk_level = vuln.risk
            if risk_level in risk_counts:
                risk_counts[risk_level] += 1
            else:
                risk_counts["Info"] += 1

        total_vulns = len(vulnerabilities)

        stats_data = [
            ["漏洞等级", "数量", "占比(%)"],
            ["超危", risk_counts["Critical"], f"{risk_counts['Critical']/total_vulns*100:.1f}%"],
            ["高危", risk_counts["High"], f"{risk_counts['High']/total_vulns*100:.1f}%"],
            ["中危", risk_counts["Medium"], f"{risk_counts['Medium']/total_vulns*100:.1f}%"],
            ["低危", risk_counts["Low"], f"{risk_counts['Low']/total_vulns*100:.1f}%"],
            ["信息", risk_counts["Info"], f"{risk_counts['Info']/total_vulns*100:.1f}%"],
            ["总计", total_vulns, "100.0%"]
        ]

        for row, data in enumerate(stats_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.alignment = self.center_alignment

    def _create_host_stats_sheet(self, ws, host_risks: Dict) -> None:
        """Create host statistics sheet"""
        host_risk_counts = {"超危": 0, "高危": 0, "中危": 0, "低危": 0, "安全": 0}
        for host, risk_info in host_risks.items():
            risk_level = risk_info['risk_level']
            if risk_level in host_risk_counts:
                host_risk_counts[risk_level] += 1

        total_hosts = len(host_risks)

        host_stats_data = [
            ["主机风险等级", "数量", "占比(%)"],
            ["超危主机", host_risk_counts["超危"], f"{host_risk_counts['超危']/total_hosts*100:.1f}%"],
            ["高危主机", host_risk_counts["高危"], f"{host_risk_counts['高危']/total_hosts*100:.1f}%"],
            ["中危主机", host_risk_counts["中危"], f"{host_risk_counts['中危']/total_hosts*100:.1f}%"],
            ["低危主机", host_risk_counts["低危"], f"{host_risk_counts['低危']/total_hosts*100:.1f}%"],
            ["安全主机", host_risk_counts["安全"], f"{host_risk_counts['安全']/total_hosts*100:.1f}%"],
            ["主机总计", total_hosts, "100.0%"]
        ]

        for row, data in enumerate(host_stats_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.alignment = self.center_alignment

    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
